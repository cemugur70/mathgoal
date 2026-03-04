import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
from datetime import datetime
from utils import get_logger, get_resource_path

logger = get_logger(__name__)

# ============== MERKEZI STIL TANIMLARI ==============
# Zebra (alternatif satır) renkleri
ZEBRA_LIGHT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_WHITE = PatternFill(fill_type=None)  # Varsayılan beyaz

# Standart border
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Ortalama hizalama
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# Koşullu format renkleri - Mevcut + Yeni
WHITE_TEXT = Font(color="FFFFFF")
GREEN_FILL = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")
BLUE_FILL = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
DARK_BLUE_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FCCF03", end_color="FCCF03", fill_type="solid")

# Alt/Üst için yumuşak renkler
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def _apply_row_style(ws, row_num, col_start=1, col_end=None):
    """
    Bir satıra zebra stili, border ve hizalama uygular.
    col_end None ise, satırdaki son dolu hücreye kadar uygular.
    """
    if col_end is None:
        col_end = ws.max_column or 50
    
    # Zebra rengi: tek satırlar açık gri, çift satırlar beyaz
    zebra_fill = ZEBRA_LIGHT if row_num % 2 == 1 else ZEBRA_WHITE
    
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        # Sadece özel bir fill uygulanmamışsa zebra uygula
        if cell.fill.fill_type is None or cell.fill.start_color.rgb in ('00000000', 'FFFFFFFF', None):
            cell.fill = zebra_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

def fractional_to_decimal(fraction_str):
    try:
        if "/" in fraction_str:
            numerator, denominator = map(float, fraction_str.split("/"))
            decimal_val = numerator / denominator + 1
        else:
            decimal_val = float(fraction_str)

        # 1) {:.2f} → "1.10"
        # 2) .replace(".", ",") → "1,10"
        result = "{:.2f}".format(decimal_val).replace(".", ",")
        
        return result

    except (ValueError, ZeroDivisionError) as e:
        # Sadece kritik hataları logla, boş string'leri değil
        if fraction_str:  # Boş değilse logla
            logger.error(f"❌ fractional_to_decimal HATA: '{fraction_str}' -> Hata: {e}")
        return ""

def _safe_get(d: dict, key: str, aliases=None, default=""):
    """dict[key] yerine güvenli erişim: alias listesi ve default destekler."""
    aliases = aliases or []
    for k in [key, *aliases]:
        if k in d:
            return d[k]
    return default

def _find_first_empty_row(ws, start_row=5, column="A"):
    """Belirtilen satırdan başlayarak bir sütundaki ilk boş satırı bulur."""
    row = start_row
    # Hücrede bir değer olduğu sürece satırı artır
    while ws[f"{column}{row}"].value not in [None, ""]:
        row += 1
    return row

def _generate_bookmaker_headers(bookmakers):
    """Dynamically generate a consistent and comprehensive set of headers for all odds."""
    headers = []
    scopes = ["", "first_half_", "second_half_"]
    
    # Define all keys and handicaps in a structured way
    bet_types = {
        "1x2": (["home", "draw", "away"], ""),
        "btts": (["yes", "no"], "btts_"),
        "dc": (["1X", "X2", "12"], "dc_"),
        "dnb": (["home", "away"], "dnb_"),
        "odd-even": (["odd", "even"], ""),
        "over-under": ([f"{i/100:.2f}".replace('.', '_') for i in range(50, 751, 25)], ""),
        "asian-handicap": ([h.replace('.', '_') for h in ["-4.5", "-4.25", "-4.0", "-3.75", "-3.5", "-3.25", "-3.0", "-2.75", "-2.5", "-2.25", "-2.0", "-1.75", "-1.5", "-1.25", "-1.0", "-0.75", "-0.5", "-0.25", "0.0", "+0.25", "+0.5", "+0.75", "+1.0", "+1.25", "+1.5", "+1.75", "+2.0", "+2.25", "+2.5", "+2.75", "+3.0", "+3.25", "+3.5", "+3.75", "+4.0", "+4.25", "+4.5"]], "ah_"),
        "ht-ft": (["1/1", "1/X", "1/2", "X/1", "X/X", "X/2", "2/1", "2/X", "2/2"], ""),
        "correct-score": ([f"{h}:{a}".replace(':', '_') for h in range(6) for a in range(6)], ""),
        "european-handicap": ([str(h) for h in ["-4", "-3", "-2", "-1", "+1", "+2", "+3", "+4"]], "eh_")
    }

    for bookmaker in bookmakers:
        for scope_prefix in scopes:
            # Simple types (1x2, btts, dc, dnb, odd-even)
            for bet_key, (outcomes, type_prefix) in bet_types.items():
                 for outcome in outcomes:
                    final_key = f"{type_prefix}{outcome}"
                    headers.extend([f"opening_{bookmaker}_{scope_prefix}{final_key}", f"{bookmaker}_{scope_prefix}{final_key}"])

            # Complex types (o/u, ah, eh)
            for bet_key in ["over-under", "asian-handicap", "european-handicap"]:
                handicaps, type_prefix = bet_types[bet_key]
                outcomes = {"over-under": ["_over", "_under"], "asian-handicap": ["_home", "_away"], "european-handicap": ["_home", "_draw", "_away"]}[bet_key]
                for handicap in handicaps:
                    for outcome in outcomes:
                        final_key = f"{type_prefix}{handicap}{outcome}"
                        headers.extend([f"opening_{bookmaker}_{scope_prefix}{final_key}", f"{bookmaker}_{scope_prefix}{final_key}"])
        
        # Non-scoped types (ht/ft, correct-score)
        for bet_key in ["ht-ft", "correct-score"]:
            handicaps, type_prefix = bet_types[bet_key]
            for handicap in handicaps:
                final_key = f"{type_prefix}{handicap}"
                headers.extend([f"opening_{bookmaker}_{final_key}", f"{bookmaker}_{final_key}"])

    return headers

def _create_headers_for_sheet(ws, sheet_name):
    """Yeni sheet için header'ları oluşturur - Renkli ve belirgin başlıklarla."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Header stilleri tanımla
    header_font = Font(bold=True, color="FFFFFF", size=10)
    basic_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Mavi
    odds_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")   # Yeşil
    special_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") # Gri
    
    # Border tanımla
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Hizalama tanımla
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Temel header'lar - Yeni sütun düzeni ile (GÜN duplicate'i düzeltildi)
    basic_headers = ["ide", "TARİH", "GÜN", "AY", "YIL", "GÜN_ADI", "SAAT", "HAFTA", "SEZON", "ÜLKE", "LİG", 
                    "EV SAHİBİ", "DEPLASMAN", "MS", "İY", "İY SONUCU", "MS SONUCU", 
                    "İY-MS", "2.5 ALT ÜST", "3.5 ÜST", "KG VAR/YOK", "İY 0.5 ALT ÜST", "İY 1.5 ALT ÜST"]
    
    # Temel header'ları yaz ve stillendir
    for i, header in enumerate(basic_headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = basic_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Sütun genişliğini ayarla
        ws.column_dimensions[cell.column_letter].width = 12
    
    # 1x2 Header'ları - Resimde görüldüğü gibi Türkçe
    x1x2_headers = [
        # FULL TIME 1X2 (21-26)
        "1",      # 21: 1 AÇILIŞ (opening)
        "1",      # 22: 1 KAPANIŞ (current)  
        "X",      # 23: X AÇILIŞ
        "X",      # 24: X KAPANIŞ
        "2",      # 25: 2 AÇILIŞ  
        "2",      # 26: 2 KAPANIŞ
        # 1st Half 1X2 (27-32)
        "1",      # 27: 1st Half 1 AÇILIŞ
        "1",      # 28: 1st Half 1 KAPANIŞ
        "X",      # 29: 1st Half X AÇILIŞ
        "X",      # 30: 1st Half X KAPANIŞ
        "2",      # 31: 1st Half 2 AÇILIŞ
        "2",      # 32: 1st Half 2 KAPANIŞ
        # 2nd Half 1X2 (33-38)
        "1",      # 33: 2nd Half 1 AÇILIŞ
        "1",      # 34: 2nd Half 1 KAPANIŞ
        "X",      # 35: 2nd Half X AÇILIŞ
        "X",      # 36: 2nd Half X KAPANIŞ
        "2",      # 37: 2nd Half 2 AÇILIŞ
        "2",      # 38: 2nd Half 2 KAPANIŞ
    ]
    
    # 1x2 header'larını yaz ve stillendir (24. sütundan başlayarak - 3 sütun kaydırıldı)
    for i, header in enumerate(x1x2_headers, 24):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = odds_fill  # 1x2 için yeşil
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 10
    
    # 2. satırda AÇILIŞ/KAPANIŞ bilgilerini yaz ve stillendir
    açılış_kapanış = [
        # FULL TIME 1X2
        "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ",
        # 1st Half 1X2  
        "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ",
        # 2nd Half 1X2
        "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ", "AÇILIŞ", "KAPANIŞ",
    ]
    
    # AÇILIŞ ve KAPANIŞ için farklı renkler
    acilis_fill = PatternFill(start_color="00CED1", end_color="00CED1", fill_type="solid")  # Turkuaz
    kapanis_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Turuncu
    
    for i, header in enumerate(açılış_kapanış, 24):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = Font(bold=True, color="000000", size=9)  # Siyah yazı
        # AÇILIŞ turkuaz, KAPANIŞ turuncu
        if header == "AÇILIŞ":
            cell.fill = acilis_fill
        else:
            cell.fill = kapanis_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # ✅ OVER/UNDER HEADER'LARI EKLE (Sütun 42-137 - 3 sütun kaydırıldı)
    col = 42
    over_under_handicaps = ["0.5", "1.5", "2.5", "3.5", "4.5", "5.5", "6.5", "7.5", 
                          "8.5", "9.5", "10.5", "11.5", "12.5", "13.5", "14.5", "15.5"]
    
    # Full Time Over/Under Headers (39-102)
    for handicap in over_under_handicaps:
        ws.cell(row=1, column=col, value=f"{handicap} ÜST")
        ws.cell(row=2, column=col, value="AÇILIŞ")
        ws.cell(row=1, column=col+1, value=f"{handicap} ÜST")
        ws.cell(row=2, column=col+1, value="KAPANIŞ")
        ws.cell(row=1, column=col+2, value=f"{handicap} ALT")
        ws.cell(row=2, column=col+2, value="AÇILIŞ")
        ws.cell(row=1, column=col+3, value=f"{handicap} ALT")
        ws.cell(row=2, column=col+3, value="KAPANIŞ")
        col += 4
    
    # 📊 RENKLİ VE BELİRGİN HEADER'LAR EKLE
    
    # Over/Under header'ları (sütun 42-137) - TURUNCU  
    ou_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    for i in range(42, 138):  # Over/Under sütunları
        cell1 = ws.cell(row=1, column=i)
        cell1.font = header_font
        cell1.fill = ou_fill
        cell1.alignment = center_alignment
        cell1.border = thin_border
        
        # Row 2: AÇILIŞ/KAPANIŞ renkleri (sütun moduna göre)
        cell2 = ws.cell(row=2, column=i)
        cell2.font = Font(bold=True, color="000000", size=9)
        # Pattern: 0-AÇILIŞ, 1-KAPANIŞ, 2-AÇILIŞ, 3-KAPANIŞ (4'lü grup)
        if (i - 42) % 2 == 0:  # AÇILIŞ sütunları (0, 2, 4...)
            cell2.fill = acilis_fill
        else:  # KAPANIŞ sütunları (1, 3, 5...)
            cell2.fill = kapanis_fill
        cell2.alignment = center_alignment
        cell2.border = thin_border
        
        ws.column_dimensions[cell1.column_letter].width = 8
    
    # Asian Handicap header'ları (sütun 138-365) - KIRMIZI  
    ah_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
    for i in range(138, 366):  # Asian Handicap sütunları
        cell1 = ws.cell(row=1, column=i, value=f"AH {i-137}")
        cell1.font = header_font
        cell1.fill = ah_fill
        cell1.alignment = center_alignment
        cell1.border = thin_border
        
        # Row 2: AÇILIŞ/KAPANIŞ renkleri (sütun moduna göre)
        cell2 = ws.cell(row=2, column=i)
        cell2.font = Font(bold=True, color="000000", size=9)
        if (i - 138) % 2 == 0:  # AÇILIŞ sütunları
            cell2.fill = acilis_fill
        else:  # KAPANIŞ sütunları
            cell2.fill = kapanis_fill
        cell2.alignment = center_alignment
        cell2.border = thin_border
        
        ws.column_dimensions[cell1.column_letter].width = 8
    
    # Diğer odds tipleri header'ları (sütun 366-461) - MOR
    other_fill = PatternFill(start_color="8A2BE2", end_color="8A2BE2", fill_type="solid")
    for i in range(366, 462):  # Diğer odds sütunları
        cell1 = ws.cell(row=1, column=i, value=f"OTHER {i-365}")
        cell1.font = header_font
        cell1.fill = other_fill
        cell1.alignment = center_alignment
        cell1.border = thin_border
        
        # Row 2: AÇILIŞ/KAPANIŞ renkleri (sütun moduna göre)
        cell2 = ws.cell(row=2, column=i)
        cell2.font = Font(bold=True, color="000000", size=9)
        if (i - 366) % 2 == 0:  # AÇILIŞ sütunları
            cell2.fill = acilis_fill
        else:  # KAPANIŞ sütunları
            cell2.fill = kapanis_fill
        cell2.alignment = center_alignment
        cell2.border = thin_border
        
        ws.column_dimensions[cell1.column_letter].width = 8
    
    # Header satırlarının yüksekliğini ayarla
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

def prepare_excel_file(excel_path):
    """Scraping başlamadan önce Excel dosyasını hazırlar - her zaman başlıkları oluşturur."""
    try:
        # Her zaman yeni Excel oluştur (template kullanma)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Tüm sheet'leri oluştur - API bookmaker isimleriyle
        from config import BOOKMAKER_MAPPING
        sheets = list(BOOKMAKER_MAPPING.keys())
        for sheet_name in sheets:
            ws = wb.create_sheet(title=sheet_name)
            _create_headers_for_sheet(ws, sheet_name)
        
        wb.save(excel_path)
        logger.info(f"✅ Excel dosyası başlıklarla oluşturuldu: {excel_path}")
    except Exception as e:
        logger.error(f"❌ Excel hazırlama hatası: {e}")



def prepare_excel_files(excel_folder):
    """Her bookmaker için ayrı Excel dosyası oluşturur."""
    from config import BOOKMAKER_MAPPING
    bookmakers = list(BOOKMAKER_MAPPING.keys())
    
    try:
        # Klasörü oluştur
        os.makedirs(excel_folder, exist_ok=True)
        
        for bookmaker in bookmakers:
            excel_path = os.path.join(excel_folder, f"{bookmaker}.xlsx")
            
            # Her bookmaker için tek sheet'li Excel oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = bookmaker
            _create_headers_for_sheet(ws, bookmaker)
            
            # Freeze panes - İlk 2 satır (başlıklar) sabit kalacak
            ws.freeze_panes = "A3"
            
            wb.save(excel_path)
            
        logger.info(f"✅ {len(bookmakers)} ayrı Excel dosyası oluşturuldu: {excel_folder}")
        return excel_folder
    except Exception as e:
        logger.error(f"❌ Excel dosyaları hazırlama hatası: {e}")
        return None

def write_to_excel(excel_path, common_data, odds_data):
    """Tek Excel dosyasına tüm bookmaker'ları yazar (eski yöntem)."""
    try:
        # Excel dosyası zaten hazırlanmış olmalı
        if not os.path.exists(excel_path):
            logger.error(f"❌ Excel dosyası bulunamadı: {excel_path}")
            return
            
        wb = load_workbook(excel_path)

        # Global stil sabitlerini kullan (dosyanın başında tanımlı)

        # Çalışılacak sayfalar - API bookmaker isimleriyle
        from config import BOOKMAKER_MAPPING
        sheets = list(BOOKMAKER_MAPPING.keys())

        for sheet in sheets:
            if sheet not in wb.sheetnames:
                logger.warning(f"⚠️ Sheet bulunamadı: {sheet} - Template'de eksik olabilir")
                continue
            ws = wb[sheet]
            
            next_row = _find_first_empty_row(ws)

            # İlk 23 kolon (temel veriler ve derived data - 3 sütun eklendi)
            for col in range(1, 24):
                header = ws.cell(row=1, column=col).value
                val = _safe_get(common_data, header)
                ws.cell(row=next_row, column=col).value = val

            under_over_keys = ["2.5 ALT ÜST", "3.5 ÜST", "KG VAR/YOK", "İY 0.5 ALT ÜST", "İY 1.5 ALT ÜST"]

            # Stil kuralları için kontrol et
            for col in range(1, 24):
                header = ws.cell(row=1, column=col).value
                val = ws.cell(row=next_row, column=col).value

                if header in ("İY SONUCU", "MS SONUCU"):
                    v = str(val)
                    if "0" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "1" in v:
                        ws.cell(row=next_row, column=col).fill = BLACK_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "2" in v:
                        ws.cell(row=next_row, column=col).fill = DARK_BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT

                if header == "İY-MS":
                    v = str(val)
                    if "İY 1/MS 2" in v or "İY 2/MS 1" in v:
                        ws.cell(row=next_row, column=col).fill = GREEN_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 0/MS 1" in v or "İY 0/MS 2" in v:
                        ws.cell(row=next_row, column=col).fill = DARK_BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 1/MS 1" in v:
                        ws.cell(row=next_row, column=col).fill = BLACK_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 2/MS 2" in v:
                        ws.cell(row=next_row, column=col).fill = BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 0/MS 0" in v or "İY 1/MS 0" in v or "İY 2/MS 0" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT

                if header in under_over_keys:
                    v = str(val)
                    if "ÜST" in v or "KG VAR" in v:
                        ws.cell(row=next_row, column=col).fill = GREEN_FILL
                    elif "ALT" in v or "KG YOK" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL

            # ---- ODDS ----
            bookmaker_name = sheet.strip() # Gizli boşlukları temizle
            row = next_row
            bookmaker = "" # for error logging
            try:
                # Sheet ismi artık doğrudan API bookmaker ismi ile aynı
                bookmaker = bookmaker_name
                
                # 1X2 FULL TIME (columns 24-29)
                home_opening = odds_data.get(f"opening_{bookmaker}_home", "")
                home_current = odds_data.get(f"{bookmaker}_home", "")
                
                converted_opening = fractional_to_decimal(home_opening)
                converted_current = fractional_to_decimal(home_current)
                
                ws.cell(row=row, column=24).value = converted_opening
                ws.cell(row=row, column=25).value = converted_current
                ws.cell(row=row, column=26).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_draw", ""))   # Col 26: X AÇILIŞ
                ws.cell(row=row, column=27).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_draw", ""))           # Col 27: X KAPANIŞ
                ws.cell(row=row, column=28).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_away", ""))   # Col 28: 2 AÇILIŞ
                ws.cell(row=row, column=29).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_away", ""))           # Col 29: 2 KAPANIŞ
                
                # 1X2 1st Half (columns 30-35) - Template structure: 1,1,X,X,2,2
                ws.cell(row=row, column=30).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_home", ""))  # Col 30: 1 AÇILIŞ
                ws.cell(row=row, column=31).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_home", ""))          # Col 31: 1 KAPANIŞ
                ws.cell(row=row, column=32).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_draw", ""))  # Col 32: X AÇILIŞ
                ws.cell(row=row, column=33).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_draw", ""))          # Col 33: X KAPANIŞ
                ws.cell(row=row, column=34).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_away", ""))  # Col 34: 2 AÇILIŞ
                ws.cell(row=row, column=35).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_away", ""))          # Col 35: 2 KAPANIŞ
                
                # 1X2 2nd Half (columns 36-41) - Template structure: 1,1,X,X,2,2 (+3 shift)
                ws.cell(row=row, column=36).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_home", "")) # Col 36: 1 AÇILIŞ
                ws.cell(row=row, column=37).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_home", ""))         # Col 37: 1 KAPANIŞ
                ws.cell(row=row, column=38).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_draw", "")) # Col 38: X AÇILIŞ
                ws.cell(row=row, column=39).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_draw", ""))         # Col 39: X KAPANIŞ
                ws.cell(row=row, column=40).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_away", "")) # Col 40: 2 AÇILIŞ
                ws.cell(row=row, column=41).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_away", ""))         # Col 41: 2 KAPANIŞ
                
                # ✅ OVER/UNDER ODDS'LARI EKLE
                
                # Over/Under 1st Half (columns 42-65) - Template: 0.5,1,1.25,1.5,2,2.5 
                # Format: handicap(OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ)
                first_half_handicaps = ["0_5", "1", "1_25", "1_5", "2", "2_5"]  # 6 handicap x 4 columns = 24 columns
                col = 42  # Template starts at column 42
                for handicap in first_half_handicaps:
                    # OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{handicap}_over", ""))      # OVER AÇILIŞ
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{handicap}_over", ""))            # OVER KAPANIŞ
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{handicap}_under", ""))    # UNDER AÇILIŞ  
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{handicap}_under", ""))          # UNDER KAPANIŞ
                    col += 4
                
                # Over/Under Full Time (columns 63-102) - Template: 0.5,1.5,2.5,2.75,3,3.5,4.5,5.5,6.5,7.5
                # Format: handicap(OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ)
                full_time_handicaps = ["0_5", "1_5", "2_5", "2_75", "3", "3_5", "4_5", "5_5", "6_5", "7_5"]  # 10 handicap x 4 columns = 40 columns
                col = 66  # Template starts at column 66 (+3 shift)
                for handicap in full_time_handicaps:
                    # OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_{handicap}_over", ""))      # OVER AÇILIŞ
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_{handicap}_over", ""))            # OVER KAPANIŞ
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_{handicap}_under", ""))    # UNDER AÇILIŞ  
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_{handicap}_under", ""))          # UNDER KAPANIŞ
                    col += 4
                
                # Over/Under 2nd Half (columns 103+) - Template: 0.5,1.5,2.5,2.75,3,3.5,4.5
                # Format: handicap(OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ) 
                second_half_handicaps = ["0_5", "1_5", "2_5", "2_75", "3", "3_5", "4_5"]  # 7 handicap x 4 columns = 28 columns 
                col = 106  # Template starts at column 106 (+3 shift)
                for handicap in second_half_handicaps:
                    # OVER AÇILIŞ, OVER KAPANIŞ, UNDER AÇILIŞ, UNDER KAPANIŞ
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_{handicap}_over", ""))      # OVER AÇILIŞ
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_{handicap}_over", ""))            # OVER KAPANIŞ
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_{handicap}_under", ""))    # UNDER AÇILIŞ
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_{handicap}_under", ""))          # UNDER KAPANIŞ
                    col += 4
                
                # ✅ ASIAN HANDICAP FULL TIME ODDS'LARI EKLE (Column 134-209)
                
                # Asian Handicap sırası (her biri 4 sütun: home_aç, home_kap, away_aç, away_kap)
                ah_handicaps = ["minus_0_25", "minus_0_5", "minus_0_75", "minus_1_0", "minus_1_25", "minus_1_5", "minus_1_75", "minus_2_0", "minus_2_25", "0_0", "0_25", "0_5", "0_75", "1_0", "1_25", "1_5", "1_75", "2_0", "2_25"]
                
                # Full Time (134'ten başlar)
                col = 134
                for handicap in ah_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_ah_{handicap}_away", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_ah_{handicap}_away", ""))
                    col += 4

                # First Half (210'dan başlar)
                col = 210
                for handicap in ah_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_ah_{handicap}_away", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_ah_{handicap}_away", ""))
                    col += 4
                
                # ✅ Diğer bet tipleri buraya eklenebilir (286+ sütunlar)
                
                # ✅ BOTH TEAMS TO SCORE ODDS'LARI EKLE (Sütun 283-294)
                
                # BTTS Full Time (sütun 283-286) - YES AÇILIŞ, YES KAPANIŞ, NO AÇILIŞ, NO KAPANIŞ
                ws.cell(row=row, column=286).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_btts_true", ""))    # YES AÇILIŞ
                ws.cell(row=row, column=287).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_btts_true", ""))          # YES KAPANIŞ
                ws.cell(row=row, column=288).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_btts_false", ""))   # NO AÇILIŞ
                ws.cell(row=row, column=289).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_btts_false", ""))           # NO KAPANIŞ
                
                # BTTS 1st Half (sütun 287-290) - YES AÇILIŞ, YES KAPANIŞ, NO AÇILIŞ, NO KAPANIŞ
                ws.cell(row=row, column=290).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_btts_true", ""))   # YES AÇILIŞ
                ws.cell(row=row, column=291).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_btts_true", ""))           # YES KAPANIŞ
                ws.cell(row=row, column=292).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_btts_false", ""))  # NO AÇILIŞ
                ws.cell(row=row, column=293).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_btts_false", ""))          # NO KAPANIŞ
                
                # BTTS 2nd Half (sütun 291-294) - YES AÇILIŞ, YES KAPANIŞ, NO AÇILIŞ, NO KAPANIŞ
                ws.cell(row=row, column=294).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_btts_true", ""))   # YES AÇILIŞ
                ws.cell(row=row, column=295).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_btts_true", ""))           # YES KAPANIŞ
                ws.cell(row=row, column=296).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_btts_false", ""))  # NO AÇILIŞ
                ws.cell(row=row, column=297).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_btts_false", ""))          # NO KAPANIŞ
                
                # ✅ DOUBLE CHANCE ODDS'LARI EKLE (Sütun 295 başlangıç)
                
                # ✅ DOUBLE CHANCE FULL TIME - Template starts at 295: 1X (295-296), 12 (297-298), X2 (299-300)
                ws.cell(row=row, column=298).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_1X", ""))      # 1X AÇILIŞ
                ws.cell(row=row, column=299).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_1X", ""))            # 1X KAPANIŞ
                ws.cell(row=row, column=300).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_12", ""))     # 12 AÇILIŞ
                ws.cell(row=row, column=301).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_12", ""))            # 12 KAPANIŞ
                ws.cell(row=row, column=302).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_X2", ""))    # X2 AÇILIŞ
                ws.cell(row=row, column=303).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_X2", ""))            # X2 KAPANIŞ
                
                # Double Chance 1st Half (sütun 301-306) - 1X AÇILIŞ, 1X KAPANIŞ, 12 AÇILIŞ, 12 KAPANIŞ, X2 AÇILIŞ, X2 KAPANIŞ
                ws.cell(row=row, column=304).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_1X", ""))   # 1X AÇILIŞ
                ws.cell(row=row, column=305).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_1X", ""))           # 1X KAPANIŞ
                ws.cell(row=row, column=306).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_12", ""))   # 12 AÇILIŞ
                ws.cell(row=row, column=307).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_12", ""))           # 12 KAPANIŞ
                ws.cell(row=row, column=308).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_X2", ""))   # X2 AÇILIŞ
                ws.cell(row=row, column=309).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_X2", ""))           # X2 KAPANIŞ
                
                # Double Chance 2nd Half (sütun 307-312) - 1X AÇILIŞ, 1X KAPANIŞ, 12 AÇILIŞ, 12 KAPANIŞ, X2 AÇILIŞ, X2 KAPANIŞ
                ws.cell(row=row, column=310).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_1X", ""))   # 1X AÇILIŞ
                ws.cell(row=row, column=311).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_1X", ""))           # 1X KAPANIŞ
                ws.cell(row=row, column=312).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_12", ""))   # 12 AÇILIŞ
                ws.cell(row=row, column=313).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_12", ""))           # 12 KAPANIŞ
                ws.cell(row=row, column=314).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_X2", ""))   # X2 AÇILIŞ
                ws.cell(row=row, column=315).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_X2", ""))           # X2 KAPANIŞ
                
                # ✅ EUROPEAN HANDICAP ODDS'LARI EKLE (Sütun 316 başlangıç)

                # European Handicap sırası (her biri 6 sütun: home_aç, home_kap, draw_aç, draw_kap, away_aç, away_kap)
                eh_handicaps = ["minus1", "plus1"]
                
                # Full Time (316'dan başlar)
                col = 316
                for handicap in eh_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+4).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_away", ""))
                    ws.cell(row=row, column=col+5).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_away", ""))
                    col += 6

                # First Half (328'den başlar)
                col = 328
                for handicap in eh_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+4).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_away", ""))
                    ws.cell(row=row, column=col+5).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_away", ""))
                    col += 6

                # ✅ DRAW NO BET ODDS'LARI EKLE (Sütun LY=315 to MJ=326)
                
                # Draw No Bet Full Time (sütun 340-343) - 1 AÇILIŞ, 1 KAPANIŞ, 2 AÇILIŞ, 2 KAPANIŞ
                ws.cell(row=row, column=340).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dnb_home", ""))   # 1 AÇILIŞ
                ws.cell(row=row, column=341).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dnb_home", ""))         # 1 KAPANIŞ
                ws.cell(row=row, column=342).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dnb_away", ""))   # 2 AÇILIŞ
                ws.cell(row=row, column=343).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dnb_away", ""))         # 2 KAPANIŞ
                
                # Draw No Bet 1st Half (sütun 344-347) - 1 AÇILIŞ, 1 KAPANIŞ, 2 AÇILIŞ, 2 KAPANIŞ
                ws.cell(row=row, column=344).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dnb_home", ""))
                ws.cell(row=row, column=345).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dnb_home", ""))
                ws.cell(row=row, column=346).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dnb_away", ""))
                ws.cell(row=row, column=347).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dnb_away", ""))
                
                # Draw No Bet 2nd Half (sütun 348-351) - 1 AÇILIŞ, 1 KAPANIŞ, 2 AÇILIŞ, 2 KAPANIŞ
                ws.cell(row=row, column=348).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dnb_home", ""))
                ws.cell(row=row, column=349).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dnb_home", ""))
                ws.cell(row=row, column=350).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dnb_away", ""))
                ws.cell(row=row, column=351).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dnb_away", ""))
                
                # ✅ CORRECT SCORE ODDS'LARI EKLE (Sütun 352 başlangıç)
                
                # Correct Score sıralaması (29 skor x 2 sütun = 58 sütun)
                correct_scores = [
                    "1_0", "2_0", "2_1", "3_0", "3_1", "3_2", "4_0", "4_1", "4_2", "4_3", "5_0", "5_1",
                    "0_0", "1_1", "2_2", "3_3", "4_4",
                    "0_1", "0_2", "1_2", "0_3", "1_3", "2_3", "0_4", "1_4", "2_4", "3_4", "0_5", "1_5"
                ]
                
                # Correct Score Full Time (sütun 352 başlangıç) - 58 sütun (352-409)
                col = 352
                for score in correct_scores:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_full_time_{score}", ""))     # AÇILIŞ
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_full_time_{score}", ""))         # KAPANIŞ
                    col += 2
                
                # Correct Score First Half (sütun 410 başlangıç) - 58 sütun (410-467)
                col = 410
                for score in correct_scores:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{score}", ""))     # AÇILIŞ
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{score}", ""))         # KAPANIŞ
                    col += 2

                # ✅ HALF TIME / FULL TIME ODDS'LARI EKLE (Sütun 468 başlangıç)
                ht_ft_outcomes = [
                    "1_1", "1_X", "1_2",
                    "X_1", "X_X", "X_2",
                    "2_1", "2_X", "2_2"
                ]

                col = 468
                for outcome in ht_ft_outcomes:
                    opening_key = f"opening_{bookmaker}_ht_ft_{outcome}"
                    closing_key = f"{bookmaker}_ht_ft_{outcome}"
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(opening_key, ""))
                    ws.cell(row=row, column=col + 1).value = fractional_to_decimal(odds_data.get(closing_key, ""))
                    col += 2

                # ✅ ODD/EVEN GOALS ODDS'LARI EKLE (Sütun 486 başlangıç)
                # Full Time (486-489)
                ws.cell(row=row, column=486).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_odd", ""))
                ws.cell(row=row, column=487).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_odd", ""))
                ws.cell(row=row, column=488).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_even", ""))
                ws.cell(row=row, column=489).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_even", ""))

                # First Half (490-493)
                ws.cell(row=row, column=490).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_odd", ""))
                ws.cell(row=row, column=491).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_odd", ""))
                ws.cell(row=row, column=492).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_even", ""))
                ws.cell(row=row, column=493).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_even", ""))

                # Second Half (494-497)
                ws.cell(row=row, column=494).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_odd", ""))
                ws.cell(row=row, column=495).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_odd", ""))
                ws.cell(row=row, column=496).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_even", ""))
                ws.cell(row=row, column=497).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_even", ""))
            except Exception as e:
                # This should catch errors for a single bookmaker and allow the loop to continue
                logger.error(f"Could not write odds for bookmaker {bookmaker} to row {row}. Error: {e}")

            # Satır stilini uygula (zebra, border, hizalama)
            _apply_row_style(ws, next_row, col_start=1, col_end=500)

        try:
            wb.save(excel_path)
        except PermissionError:
            logger.error(f"❌ KAYDETME BAŞARISIZ: Excel dosyası '{excel_path}' başka bir program tarafından kullanılıyor olabilir. Lütfen dosyayı kapatıp tekrar deneyin.")

    except Exception as e:
        import traceback
        sheet_name = sheet if "sheet" in locals() else "?"
        last_header = header if "header" in locals() else "?"
        logger.error(f"[write_to_excel ERROR] sheet={sheet_name} header={last_header} err={repr(e)}")
        raise


def write_to_excel_separate(excel_folder, common_data, odds_data):
    """Her bookmaker için ayrı Excel dosyasına yazar."""
    from config import BOOKMAKER_MAPPING
    bookmakers = list(BOOKMAKER_MAPPING.keys())
    
    # Global stil sabitlerini kullan (dosyanın başında tanımlı)
    under_over_keys = ["2.5 ALT ÜST", "3.5 ÜST", "KG VAR/YOK", "İY 0.5 ALT ÜST", "İY 1.5 ALT ÜST"]
    
    for bookmaker in bookmakers:
        excel_path = os.path.join(excel_folder, f"{bookmaker}.xlsx")
        
        try:
            if not os.path.exists(excel_path):
                logger.error(f"❌ Excel dosyası bulunamadı: {excel_path}")
                continue
                
            wb = load_workbook(excel_path)
            ws = wb.active
            
            next_row = _find_first_empty_row(ws)

            # İlk 23 kolon (temel veriler ve derived data)
            for col in range(1, 24):
                header = ws.cell(row=1, column=col).value
                val = _safe_get(common_data, header)
                ws.cell(row=next_row, column=col).value = val

            # Stil kuralları için kontrol et
            for col in range(1, 24):
                header = ws.cell(row=1, column=col).value
                val = ws.cell(row=next_row, column=col).value

                if header in ("İY SONUCU", "MS SONUCU"):
                    v = str(val)
                    if "0" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "1" in v:
                        ws.cell(row=next_row, column=col).fill = BLACK_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "2" in v:
                        ws.cell(row=next_row, column=col).fill = DARK_BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT

                if header == "İY-MS":
                    v = str(val)
                    if "İY 1/MS 2" in v or "İY 2/MS 1" in v:
                        ws.cell(row=next_row, column=col).fill = GREEN_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 0/MS 1" in v or "İY 0/MS 2" in v:
                        ws.cell(row=next_row, column=col).fill = DARK_BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 1/MS 1" in v:
                        ws.cell(row=next_row, column=col).fill = BLACK_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 2/MS 2" in v:
                        ws.cell(row=next_row, column=col).fill = BLUE_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT
                    elif "İY 0/MS 0" in v or "İY 1/MS 0" in v or "İY 2/MS 0" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL
                        ws.cell(row=next_row, column=col).font = WHITE_TEXT

                if header in under_over_keys:
                    v = str(val)
                    if "ÜST" in v or "KG VAR" in v:
                        ws.cell(row=next_row, column=col).fill = GREEN_FILL
                    elif "ALT" in v or "KG YOK" in v:
                        ws.cell(row=next_row, column=col).fill = RED_FILL

            # ---- ODDS ----
            row = next_row
            try:
                # 1X2 FULL TIME (columns 24-29)
                ws.cell(row=row, column=24).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_home", ""))
                ws.cell(row=row, column=25).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_home", ""))
                ws.cell(row=row, column=26).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_draw", ""))
                ws.cell(row=row, column=27).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_draw", ""))
                ws.cell(row=row, column=28).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_away", ""))
                ws.cell(row=row, column=29).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_away", ""))
                
                # 1X2 1st Half (columns 30-35)
                ws.cell(row=row, column=30).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_home", ""))
                ws.cell(row=row, column=31).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_home", ""))
                ws.cell(row=row, column=32).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_draw", ""))
                ws.cell(row=row, column=33).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_draw", ""))
                ws.cell(row=row, column=34).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_away", ""))
                ws.cell(row=row, column=35).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_away", ""))
                
                # 1X2 2nd Half (columns 36-41)
                ws.cell(row=row, column=36).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_home", ""))
                ws.cell(row=row, column=37).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_home", ""))
                ws.cell(row=row, column=38).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_draw", ""))
                ws.cell(row=row, column=39).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_draw", ""))
                ws.cell(row=row, column=40).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_away", ""))
                ws.cell(row=row, column=41).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_away", ""))
                
                # Over/Under 1st Half (columns 42-65)
                first_half_handicaps = ["0_5", "1", "1_25", "1_5", "2", "2_5"]
                col = 42
                for handicap in first_half_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{handicap}_over", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{handicap}_over", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{handicap}_under", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{handicap}_under", ""))
                    col += 4
                
                # Over/Under Full Time (columns 66-105)
                full_time_handicaps = ["0_5", "1_5", "2_5", "2_75", "3", "3_5", "4_5", "5_5", "6_5", "7_5"]
                col = 66
                for handicap in full_time_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_{handicap}_over", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_{handicap}_over", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_{handicap}_under", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_{handicap}_under", ""))
                    col += 4
                
                # Over/Under 2nd Half (columns 106+)
                second_half_handicaps = ["0_5", "1_5", "2_5", "2_75", "3", "3_5", "4_5"]
                col = 106
                for handicap in second_half_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_{handicap}_over", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_{handicap}_over", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_{handicap}_under", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_{handicap}_under", ""))
                    col += 4
                
                # Asian Handicap Full Time (Column 134-209)
                ah_handicaps = ["minus_0_25", "minus_0_5", "minus_0_75", "minus_1_0", "minus_1_25", "minus_1_5", "minus_1_75", "minus_2_0", "minus_2_25", "0_0", "0_25", "0_5", "0_75", "1_0", "1_25", "1_5", "1_75", "2_0", "2_25"]
                col = 134
                for handicap in ah_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_ah_{handicap}_away", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_ah_{handicap}_away", ""))
                    col += 4

                # Asian Handicap First Half (210+)
                col = 210
                for handicap in ah_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_ah_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_ah_{handicap}_away", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_ah_{handicap}_away", ""))
                    col += 4
                
                # BTTS Full Time (286-289)
                ws.cell(row=row, column=286).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_btts_true", ""))
                ws.cell(row=row, column=287).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_btts_true", ""))
                ws.cell(row=row, column=288).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_btts_false", ""))
                ws.cell(row=row, column=289).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_btts_false", ""))
                
                # BTTS 1st Half (290-293)
                ws.cell(row=row, column=290).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_btts_true", ""))
                ws.cell(row=row, column=291).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_btts_true", ""))
                ws.cell(row=row, column=292).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_btts_false", ""))
                ws.cell(row=row, column=293).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_btts_false", ""))
                
                # BTTS 2nd Half (294-297)
                ws.cell(row=row, column=294).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_btts_true", ""))
                ws.cell(row=row, column=295).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_btts_true", ""))
                ws.cell(row=row, column=296).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_btts_false", ""))
                ws.cell(row=row, column=297).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_btts_false", ""))
                
                # Double Chance Full Time (298-303)
                ws.cell(row=row, column=298).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_1X", ""))
                ws.cell(row=row, column=299).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_1X", ""))
                ws.cell(row=row, column=300).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_12", ""))
                ws.cell(row=row, column=301).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_12", ""))
                ws.cell(row=row, column=302).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dc_X2", ""))
                ws.cell(row=row, column=303).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dc_X2", ""))
                
                # Double Chance 1st Half (304-309)
                ws.cell(row=row, column=304).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_1X", ""))
                ws.cell(row=row, column=305).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_1X", ""))
                ws.cell(row=row, column=306).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_12", ""))
                ws.cell(row=row, column=307).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_12", ""))
                ws.cell(row=row, column=308).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dc_X2", ""))
                ws.cell(row=row, column=309).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dc_X2", ""))
                
                # Double Chance 2nd Half (310-315)
                ws.cell(row=row, column=310).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_1X", ""))
                ws.cell(row=row, column=311).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_1X", ""))
                ws.cell(row=row, column=312).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_12", ""))
                ws.cell(row=row, column=313).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_12", ""))
                ws.cell(row=row, column=314).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dc_X2", ""))
                ws.cell(row=row, column=315).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dc_X2", ""))
                
                # European Handicap Full Time (316-327)
                eh_handicaps = ["minus1", "plus1"]
                col = 316
                for handicap in eh_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+4).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_eh_{handicap}_away", ""))
                    ws.cell(row=row, column=col+5).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_eh_{handicap}_away", ""))
                    col += 6

                # European Handicap First Half (328-339)
                col = 328
                for handicap in eh_handicaps:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_home", ""))
                    ws.cell(row=row, column=col+2).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+3).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_draw", ""))
                    ws.cell(row=row, column=col+4).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_eh_{handicap}_away", ""))
                    ws.cell(row=row, column=col+5).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_eh_{handicap}_away", ""))
                    col += 6
                
                # Draw No Bet Full Time (340-343)
                ws.cell(row=row, column=340).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dnb_home", ""))
                ws.cell(row=row, column=341).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dnb_home", ""))
                ws.cell(row=row, column=342).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_dnb_away", ""))
                ws.cell(row=row, column=343).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_dnb_away", ""))
                
                # Draw No Bet 1st Half (344-347)
                ws.cell(row=row, column=344).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dnb_home", ""))
                ws.cell(row=row, column=345).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dnb_home", ""))
                ws.cell(row=row, column=346).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_dnb_away", ""))
                ws.cell(row=row, column=347).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_dnb_away", ""))
                
                # Draw No Bet 2nd Half (348-351)
                ws.cell(row=row, column=348).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dnb_home", ""))
                ws.cell(row=row, column=349).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dnb_home", ""))
                ws.cell(row=row, column=350).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_dnb_away", ""))
                ws.cell(row=row, column=351).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_dnb_away", ""))
                
                # Correct Score Full Time (352-409)
                correct_scores = [
                    "1_0", "2_0", "2_1", "3_0", "3_1", "3_2", "4_0", "4_1", "4_2", "4_3", "5_0", "5_1",
                    "0_0", "1_1", "2_2", "3_3", "4_4",
                    "0_1", "0_2", "1_2", "0_3", "1_3", "2_3", "0_4", "1_4", "2_4", "3_4", "0_5", "1_5"
                ]
                col = 352
                for score in correct_scores:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_full_time_{score}", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_full_time_{score}", ""))
                    col += 2
                
                # Correct Score First Half (410-467)
                col = 410
                for score in correct_scores:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_{score}", ""))
                    ws.cell(row=row, column=col+1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_{score}", ""))
                    col += 2

                # Half Time / Full Time (468-485)
                ht_ft_outcomes = ["1_1", "1_X", "1_2", "X_1", "X_X", "X_2", "2_1", "2_X", "2_2"]
                col = 468
                for outcome in ht_ft_outcomes:
                    ws.cell(row=row, column=col).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_ht_ft_{outcome}", ""))
                    ws.cell(row=row, column=col + 1).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_ht_ft_{outcome}", ""))
                    col += 2

                # Odd/Even Full Time (486-489)
                ws.cell(row=row, column=486).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_odd", ""))
                ws.cell(row=row, column=487).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_odd", ""))
                ws.cell(row=row, column=488).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_even", ""))
                ws.cell(row=row, column=489).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_even", ""))

                # Odd/Even First Half (490-493)
                ws.cell(row=row, column=490).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_odd", ""))
                ws.cell(row=row, column=491).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_odd", ""))
                ws.cell(row=row, column=492).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_first_half_even", ""))
                ws.cell(row=row, column=493).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_first_half_even", ""))

                # Odd/Even Second Half (494-497)
                ws.cell(row=row, column=494).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_odd", ""))
                ws.cell(row=row, column=495).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_odd", ""))
                ws.cell(row=row, column=496).value = fractional_to_decimal(odds_data.get(f"opening_{bookmaker}_second_half_even", ""))
                ws.cell(row=row, column=497).value = fractional_to_decimal(odds_data.get(f"{bookmaker}_second_half_even", ""))
                
            except Exception as e:
                logger.error(f"Could not write odds for bookmaker {bookmaker} to row {row}. Error: {e}")

            # Satır stilini uygula (zebra, border, hizalama)
            _apply_row_style(ws, next_row, col_start=1, col_end=500)

            try:
                wb.save(excel_path)
            except PermissionError:
                logger.error(f"❌ KAYDETME BAŞARISIZ: Excel dosyası '{excel_path}' başka bir program tarafından kullanılıyor olabilir.")

        except Exception as e:
            import traceback
            logger.error(f"[write_to_excel_separate ERROR] bookmaker={bookmaker} err={repr(e)}")


def sort_excel_files(excel_folder):
    """Klasördeki tüm Excel dosyalarını sıralar."""
    bookmakers = ["bet365", "BetMGM", "Betfred", "Unibetuk", "Betway", "Midnite", "Ladbrokes", "Betfair", "7Bet"]
    
    for bookmaker in bookmakers:
        excel_path = os.path.join(excel_folder, f"{bookmaker}.xlsx")
        if os.path.exists(excel_path):
            sort_excel_file(excel_path)
    
    logger.info(f"✅ Tüm Excel dosyaları sıralandı: {excel_folder}")

def sort_excel_file(excel_path):
    """Excel dosyasındaki verileri tarihe göre (Yeniden Eskiye) sıralar ve stilleri yeniden uygular."""
    try:
        if not os.path.exists(excel_path):
            logger.error(f"❌ Sıralama için Excel dosyası bulunamadı: {excel_path}")
            return

        wb = load_workbook(excel_path)
        
        # Global stil sabitlerini kullan (dosyanın başında tanımlı)
        under_over_keys = ["2.5 ALT ÜST", "3.5 ÜST", "KG VAR/YOK", "İY 0.5 ALT ÜST", "İY 1.5 ALT ÜST"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Veri satırlarını al (5. satırdan itibaren)
            data_rows = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if row[0] is None: # Boş satır kontrolü
                    continue
                data_rows.append(row)
            
            if not data_rows:
                continue
                
            # Sıralama fonksiyonu
            def parse_datetime(row):
                try:
                    date_str = row[1] # TARİH (Col 2) -> Index 1
                    time_str = row[6] # SAAT (Col 7) -> Index 6
                    
                    # Yıl kontrolü (Eğer TARİH sadece gün.ay ise, YIL sütununu kullan)
                    if date_str and str(date_str).count('.') == 2:
                        dt_str = f"{date_str} {time_str}"
                        return datetime.strptime(dt_str, "%d.%m.%Y %H:%M")
                    else:
                        # Fallback: YIL sütununu (Col 5 -> Index 4) kullan
                        year = row[4]
                        dt_str = f"{date_str}.{year} {time_str}"
                        return datetime.strptime(dt_str, "%d.%m.%Y %H:%M")
                except Exception:
                    return datetime.min # Hata durumunda en sona at (veya en başa)

            # Yeniden eskiye sırala (reverse=True)
            data_rows.sort(key=parse_datetime, reverse=True)
            
            # Mevcut verileri temizle (sadece değerleri değil, satırları silip yeniden oluşturmak daha temiz olabilir ama formatı korumak zor)
            # Bu yüzden hücre değerlerini güncelleyeceğiz.
            
            # Önce tüm veri alanını temizle (Values only)
            # Ancak satır sayısı değişmeyeceği için (sadece sıra değişiyor), direkt üzerine yazabiliriz.
            
            for i, row_values in enumerate(data_rows):
                row_idx = 5 + i
                for col_idx, value in enumerate(row_values):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    cell.value = value
                    
                    # --- Stil Kurallarını Yeniden Uygula (Sütun 1-23) ---
                    # Sadece ilk 23 sütun için stil var
                    if col_idx < 23:
                        # Header ismini bul (Row 1)
                        header = ws.cell(row=1, column=col_idx + 1).value
                        
                        # Temizle (Varsayılan stil)
                        cell.fill = PatternFill(fill_type=None)
                        cell.font = Font(color="000000")

                        if header in ("İY SONUCU", "MS SONUCU"):
                            v = str(value)
                            if "0" in v:
                                cell.fill = RED_FILL
                                cell.font = WHITE_TEXT
                            elif "1" in v:
                                cell.fill = BLACK_FILL
                                cell.font = WHITE_TEXT
                            elif "2" in v:
                                cell.fill = DARK_BLUE_FILL
                                cell.font = WHITE_TEXT

                        elif header == "İY-MS":
                            v = str(value)
                            if "İY 1/MS 2" in v or "İY 2/MS 1" in v:
                                cell.fill = GREEN_FILL
                                cell.font = WHITE_TEXT
                            elif "İY 0/MS 1" in v or "İY 0/MS 2" in v:
                                cell.fill = DARK_BLUE_FILL
                                cell.font = WHITE_TEXT
                            elif "İY 1/MS 1" in v:
                                cell.fill = BLACK_FILL
                                cell.font = WHITE_TEXT
                            elif "İY 2/MS 2" in v:
                                cell.fill = BLUE_FILL
                                cell.font = WHITE_TEXT
                            elif "İY 0/MS 0" in v or "İY 1/MS 0" in v or "İY 2/MS 0" in v:
                                cell.fill = RED_FILL
                                cell.font = WHITE_TEXT

                        elif header in under_over_keys:
                            v = str(value)
                            if "ÜST" in v or "KG VAR" in v:
                                cell.fill = GREEN_FILL
                            elif "ALT" in v or "KG YOK" in v:
                                cell.fill = RED_FILL
                
                # Satır stilini uygula (zebra, border, hizalama)
                _apply_row_style(ws, row_idx, col_start=1, col_end=500)

        wb.save(excel_path)

    except Exception as e:
        logger.error(f"❌ Excel sıralama hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
