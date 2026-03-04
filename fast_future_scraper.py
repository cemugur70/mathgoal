"""
FAST FUTURE SCRAPER - For upcoming matches (Yeni Maçlar)
Uses EXACT SAME Excel format as fast_scraper.py
Simply wraps fast_scraper.run_threaded_scraper and marks scores as '-'
"""
from fast_scraper import run_threaded_scraper as _run_season_scraper, scrape_match_data
from datetime import datetime
import threading

# Thread-safe results list (for direct scraping)
RESULTS = []
RESULTS_LOCK = threading.Lock()
PROGRESS = {"done": 0, "total": 0}
PROGRESS_LOCK = threading.Lock()


def scrape_future_match_data(match_id: str, bookmakers: list, bet_types: dict, logger) -> dict:
    """Scrape data for an UPCOMING match - reuse fast_scraper logic but mark scores as '-'"""
    result = scrape_match_data(match_id, bookmakers, bet_types, logger)
    
    if result:
        # Override scores with '-' since match hasn't been played yet
        result['MS'] = '-'
        result['İY'] = '-'
        result['İY SONUCU'] = '-'
        result['MS SONUCU'] = '-'
        result['İY-MS'] = '-'
        result['2.5 ALT ÜST'] = '-'
        result['3.5 ÜST'] = '-'
        result['KG VAR/YOK'] = '-'
        result['İY 0.5 ALT ÜST'] = '-'
        result['İY 1.5 ALT ÜST'] = '-'
    
    return result


def run_future_scraper(match_ids: list, bookmakers: list, bet_types: dict, excel_folder: str, logger, max_workers: int = 20, datetime_map: dict = None, odds_option: str = "both"):
    """
    Fast scraper for FUTURE matches.
    Uses EXACT SAME Excel format as season scraper.
    Only difference: score fields are set to '-' after scraping.
    excel_folder: Folder path where separate Excel files per bookmaker are created
    odds_option: "both" (tümü), "opening" (sadece açılış), "closing" (sadece kapanış)
    """
    if datetime_map is None:
        datetime_map = {}
    
    logger.info(f"🚀 Yeni maçlar taranıyor: {len(match_ids)} maç")
    logger.info(f"📊 Oran seçeneği: {odds_option}")
    
    # Use the EXACT SAME scraper as season data
    # This ensures 100% identical Excel format
    failed, results = _run_season_scraper(
        match_ids, bookmakers, bet_types, excel_folder, logger, max_workers, datetime_map, odds_option
    )
    
    # Now we need to re-process all Excel files in the folder to mark scores as '-'
    import os
    import openpyxl
    
    # Score columns to mark as '-'
    score_cols = ['İY', 'MS', 'İY SONUCU', 'MS SONUCU', 'İY-MS', 
                  '2.5 ALT ÜST', '3.5 ÜST', 'KG VAR/YOK', 
                  'İY 0.5 ALT ÜST', 'İY 1.5 ALT ÜST']
    
    try:
        # Process all Excel files in the folder
        for xlsx_file in os.listdir(excel_folder):
            if xlsx_file.endswith('.xlsx'):
                excel_path = os.path.join(excel_folder, xlsx_file)
                try:
                    wb = openpyxl.load_workbook(excel_path)
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # Find header row (row 1 for new format)
                        header_row = 1
                        headers = {}
                        for col_idx, cell in enumerate(ws[header_row], 1):
                            if cell.value:
                                headers[cell.value] = col_idx
                        
                        # Find columns to update
                        cols_to_update = []
                        for score_col in score_cols:
                            if score_col in headers:
                                cols_to_update.append(headers[score_col])
                        
                        # Update data rows (starting from row 3 - after header and AÇILIŞ/KAPANIŞ row)
                        for row_idx in range(3, ws.max_row + 1):
                            for col_idx in cols_to_update:
                                ws.cell(row=row_idx, column=col_idx, value='-')
                    
                    wb.save(excel_path)
                    wb.close()
                except Exception as e:
                    logger.warning(f"Skor güncelleme hatası ({xlsx_file}): {e}")
        
        logger.info(f"✅ Skorlar '-' olarak işaretlendi")
        
    except Exception as e:
        logger.warning(f"Klasör işleme hatası: {e}")
    
    return failed, results
